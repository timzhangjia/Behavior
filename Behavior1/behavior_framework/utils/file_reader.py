"""
File Reader Utility Module - Provides file reading functionality for YAML, Excel, etc.
"""

import yaml
import json
import os
from typing import Dict, Any, Optional, List
from pathlib import Path
from openpyxl import load_workbook
from ..utils.logger import Logger


class FileReader:
    """File reader utility class"""
    
    def __init__(self):
        self.logger = Logger()
    
    def read_yaml(self, file_path: str) -> Dict[str, Any]:
        """
        Read YAML file
        
        Args:
            file_path: YAML file path
            
        Returns:
            Dictionary format data
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
                self.logger.debug(f"Successfully read YAML file: {file_path}")
                return data or {}
        except Exception as e:
            self.logger.error(f"Failed to read YAML file: {file_path}, error: {str(e)}")
            raise
    
    def read_json(self, file_path: str) -> Dict[str, Any]:
        """
        Read JSON file
        
        Args:
            file_path: JSON file path
            
        Returns:
            Dictionary format data
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.logger.debug(f"Successfully read JSON file: {file_path}")
                return data
        except Exception as e:
            self.logger.error(f"Failed to read JSON file: {file_path}, error: {str(e)}")
            raise
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Read Excel file
        
        Args:
            file_path: Excel file path
            sheet_name: Worksheet name, if None reads the first worksheet
            
        Returns:
            List format data, each element is a dictionary of row data
        """
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Select worksheet
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            
            # Read headers
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            # Read data
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    data.append(row_data)
            
            workbook.close()
            self.logger.debug(f"Successfully read Excel file: {file_path}, {len(data)} rows of data")
            return data
        except Exception as e:
            self.logger.error(f"Failed to read Excel file: {file_path}, error: {str(e)}")
            raise
    
    def get_element_locator(self, page_name: str, element_name: str, base_dir: str = "data/elements") -> Dict[str, str]:
        """
        Get element locator information from YAML file
        
        Args:
            page_name: Page name
            element_name: Element name
            base_dir: Base directory for element location files
            
        Returns:
            Dictionary containing type and value, e.g.: {"type": "id", "value": "username"}
        """
        try:
            # Find corresponding yaml file
            yaml_file = Path(base_dir) / f"{page_name}.yaml"
            
            # If specific page yaml file doesn't exist, try reading common.yaml
            if not yaml_file.exists():
                yaml_file = Path(base_dir) / "common.yaml"
            
            if not yaml_file.exists():
                self.logger.warning(f"Element location file does not exist: {yaml_file}")
                return {}
            
            # Read yaml file
            data = self.read_yaml(str(yaml_file))
            
            # Get element location information
            if page_name in data and element_name in data[page_name]:
                return data[page_name][element_name]
            elif element_name in data:
                # If element found directly in common.yaml
                return data[element_name]
            else:
                self.logger.warning(f"Element location information not found: {page_name}.{element_name}")
                return {}
        except Exception as e:
            self.logger.error(f"Failed to get element location information: {str(e)}")
            return {}


class ElementLocator:
    """Element locator utility class - encapsulates methods for reading element locations from yaml"""
    
    def __init__(self, base_dir: str = "data/elements"):
        self.file_reader = FileReader()
        self.base_dir = base_dir
    
    def get_locator(self, page_name: str, element_name: str) -> Dict[str, str]:
        """
        Get element location information
        
        Args:
            page_name: Page name
            element_name: Element name
            
        Returns:
            Dictionary containing type and value
        """
        return self.file_reader.get_element_locator(page_name, element_name, self.base_dir)
    
    def get_selector(self, page_name: str, element_name: str) -> str:
        """
        Get element selector string
        
        Args:
            page_name: Page name
            element_name: Element name
            
        Returns:
            Selector string
        """
        locator = self.get_locator(page_name, element_name)
        if not locator:
            return ""
        
        locator_type = locator.get("type", "").lower()
        value = locator.get("value", "")
        
        if locator_type == "id":
            return f"#{value}"
        elif locator_type == "class":
            return f".{value}"
        elif locator_type == "css":
            return value
        elif locator_type == "xpath":
            return value
        elif locator_type == "name":
            return f"[name='{value}']"
        else:
            return value
